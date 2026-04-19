"""대회 프로그램 (진행표) CRUD + 엑셀 임포트."""

import io
import re
from datetime import date
from typing import Optional, List
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from pydantic import BaseModel, Field
from sqlalchemy import select, and_
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.mappings import (
    EVENT_TYPE_MAP,
    DIVISION_NORMALIZE,
    GENDER_MAP,
    TEAM_EVENT_CODES,
)
from app.models import Program, Participant
from app.schemas import APIResponse

router = APIRouter(prefix="/events/{event_id}/programs", tags=["programs"])


class ProgramCreate(BaseModel):
    order: int = 0
    division: str = Field(..., min_length=1, max_length=50)
    event_type: str = Field(..., min_length=1, max_length=50)
    event_code: Optional[str] = None
    round: str = "본선"
    heat_duration_sec: int = 30
    participants_per_heat: int = 1
    note: Optional[str] = None


class ProgramUpdate(BaseModel):
    order: Optional[int] = None
    division: Optional[str] = None
    event_type: Optional[str] = None
    event_code: Optional[str] = None
    round: Optional[str] = None
    heat_duration_sec: Optional[int] = None
    participants_per_heat: Optional[int] = None
    note: Optional[str] = None


class ProgramResponse(BaseModel):
    id: UUID
    event_id: UUID
    order: int
    division: str
    event_type: str
    event_code: Optional[str] = None
    round: str
    heat_duration_sec: int
    participants_per_heat: int
    note: Optional[str] = None
    competition_date: Optional[date] = None
    heat_assignments: Optional[list] = None

    class Config:
        from_attributes = True


class HeatAssignmentUpdate(BaseModel):
    heat_assignments: list  # [{heat_number: int, participant_ids: [str, ...]}]


class ProgramListResponse(BaseModel):
    items: List[ProgramResponse]
    total: int


@router.post("", response_model=APIResponse[ProgramResponse])
async def create_program(
    event_id: UUID,
    body: ProgramCreate,
    db: AsyncSession = Depends(get_db),
):
    """프로그램 항목 추가."""
    program = Program(
        event_id=event_id,
        order=body.order,
        division=body.division,
        event_type=body.event_type,
        event_code=body.event_code,
        round=body.round,
        heat_duration_sec=body.heat_duration_sec,
        participants_per_heat=body.participants_per_heat,
        note=body.note,
    )
    db.add(program)
    await db.commit()
    await db.refresh(program)
    return APIResponse(data=ProgramResponse.model_validate(program))


@router.get("", response_model=APIResponse[ProgramListResponse])
async def list_programs(
    event_id: UUID,
    db: AsyncSession = Depends(get_db),
):
    """이벤트의 프로그램 목록 (순번 정렬)."""
    q = (
        select(Program)
        .where(Program.event_id == event_id)
        .order_by(Program.order)
    )
    result = await db.execute(q)
    items = result.scalars().all()
    return APIResponse(
        data=ProgramListResponse(
            items=[ProgramResponse.model_validate(p) for p in items],
            total=len(items),
        )
    )


@router.patch("/{program_id}", response_model=APIResponse[ProgramResponse])
async def update_program(
    event_id: UUID,
    program_id: UUID,
    body: ProgramUpdate,
    db: AsyncSession = Depends(get_db),
):
    """프로그램 항목 수정."""
    result = await db.execute(
        select(Program).where(Program.id == program_id, Program.event_id == event_id)
    )
    program = result.scalar_one_or_none()
    if not program:
        raise HTTPException(status_code=404, detail="프로그램을 찾을 수 없습니다")

    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(program, field, value)
    await db.commit()
    await db.refresh(program)
    return APIResponse(data=ProgramResponse.model_validate(program))


@router.put("/{program_id}/assignments", response_model=APIResponse[ProgramResponse])
async def update_heat_assignments(
    event_id: UUID,
    program_id: UUID,
    body: HeatAssignmentUpdate,
    db: AsyncSession = Depends(get_db),
):
    """히트 배정표 저장."""
    result = await db.execute(
        select(Program).where(Program.id == program_id, Program.event_id == event_id)
    )
    program = result.scalar_one_or_none()
    if not program:
        raise HTTPException(status_code=404, detail="프로그램을 찾을 수 없습니다")

    program.heat_assignments = body.heat_assignments
    await db.commit()
    await db.refresh(program)
    return APIResponse(data=ProgramResponse.model_validate(program))


@router.delete("/{program_id}", response_model=APIResponse[dict])
async def delete_program(
    event_id: UUID,
    program_id: UUID,
    db: AsyncSession = Depends(get_db),
):
    """프로그램 항목 삭제."""
    result = await db.execute(
        select(Program).where(Program.id == program_id, Program.event_id == event_id)
    )
    program = result.scalar_one_or_none()
    if not program:
        raise HTTPException(status_code=404, detail="프로그램을 찾을 수 없습니다")
    await db.delete(program)
    await db.commit()
    return APIResponse(data={"deleted": True})


# ── 엑셀 임포트 ────────────────────────────────────────────────


def _parse_division(raw: str):
    """'M U9', 'F 12-15', 'Mixed U16', 'X 16+' 등을 (gender_code, age_div) 로 분리."""
    raw = raw.strip()
    parts = raw.split(None, 1)
    if len(parts) == 2:
        gender_code, age_raw = parts
    else:
        gender_code, age_raw = "X", parts[0]
    gender = GENDER_MAP.get(gender_code, "혼성")
    age_div = DIVISION_NORMALIZE.get(age_raw, age_raw)
    return gender, age_div


def _clean_phone(phone) -> str:
    if phone is None:
        return ""
    return re.sub(r"[^0-9+]", "", str(phone))


def _parse_date_from_sheet(sheet_name: str):
    """시트 이름에서 날짜 패턴을 추출.

    지원 형식:
      - 'YYYYMMDD'  (예: '20260502')
      - '스피드_5월3일' (예: 한글 날짜)
    """
    import datetime as _dt

    # YYYYMMDD 형식 (예: 20260502)
    m = re.match(r"^(\d{4})(\d{2})(\d{2})$", sheet_name.strip())
    if m:
        try:
            return _dt.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass

    # 한글 날짜 형식 (예: 스피드_5월3일)
    m = re.search(r"(\d{1,2})월(\d{1,2})일", sheet_name)
    if m:
        month, day = int(m.group(1)), int(m.group(2))
        year = _dt.date.today().year
        try:
            return _dt.date(year, month, day)
        except ValueError:
            pass
    return None


async def _find_or_create_participant(
    db: AsyncSession,
    event_id: UUID,
    name: str,
    phone: str,
    email: str | None,
    team: str | None,
    category: str | None,
) -> Participant:
    """참가자 생성. 중복 허용 — 같은 이름이라도 별개 참가 건으로 등록.

    동일 (이름 + 연락처 + 카테고리) 조합만 중복 방지 (같은 행 재임포트 시).
    """
    if phone:
        stmt = select(Participant).where(
            and_(
                Participant.event_id == event_id,
                Participant.name == name,
                Participant.phone == phone,
                Participant.category == category,
            )
        )
        result = await db.execute(stmt)
        p = result.scalars().first()
        if p:
            return p

    p = Participant(
        event_id=event_id,
        name=name.strip(),
        phone=phone,
        email=email or None,
        team=team or None,
        category=category,
    )
    db.add(p)
    await db.flush()
    return p


async def _find_or_create_program(
    db: AsyncSession,
    event_id: UUID,
    event_code: str,
    division_raw: str,
    competition_date,
    order_counter: dict,
) -> Program:
    """event_code + division 원본 조합으로 기존 프로그램 찾거나 생성."""
    division_label = division_raw
    evt_info = EVENT_TYPE_MAP.get(event_code, {})
    event_type_name = evt_info.get("name", event_code)

    stmt = select(Program).where(
        and_(
            Program.event_id == event_id,
            Program.event_code == event_code,
            Program.division == division_label,
            Program.competition_date == competition_date,
        )
    )
    result = await db.execute(stmt)
    prog = result.scalar_one_or_none()
    if prog:
        return prog

    order_counter["n"] += 1
    prog = Program(
        event_id=event_id,
        order=order_counter["n"],
        division=division_label,
        event_type=event_type_name,
        event_code=event_code,
        round="본선",
        heat_duration_sec=evt_info.get("duration", 30),
        participants_per_heat=evt_info.get("per_heat", 1),
        competition_date=competition_date,
        heat_assignments=[],
    )
    db.add(prog)
    await db.flush()
    return prog


class ImportSummary(BaseModel):
    programs_created: int
    participants_imported: int
    heats_assigned: int
    sheets_processed: list[str]
    errors: list[str]


@router.post("/import-excel", response_model=APIResponse[ImportSummary])
async def import_excel(
    event_id: UUID,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """엑셀 파일을 파싱하여 프로그램·참가자·히트 배정을 일괄 생성."""
    import openpyxl

    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="xlsx 파일만 지원합니다")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)

    # Re-import: 기존 데이터 삭제 (full reset)
    existing_participants = await db.execute(
        select(Participant).where(Participant.event_id == event_id)
    )
    for p in existing_participants.scalars().all():
        await db.delete(p)

    existing_programs = await db.execute(
        select(Program).where(Program.event_id == event_id)
    )
    for prog in existing_programs.scalars().all():
        await db.delete(prog)
    await db.flush()

    # order 는 0부터 다시 시작
    order_counter = {"n": 0}

    total_programs = 0
    total_participants = 0
    total_heats = 0
    sheets_processed = []
    errors: list[str] = []

    # 프로그램별 히트 배정 누적용
    program_heats: dict[str, dict] = {}  # prog.id -> {heat_key -> {heat_number, station, participant_ids}}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if len(rows) < 3:
            continue  # 헤더 + 데이터가 없으면 스킵

        sheets_processed.append(sheet_name)
        competition_date = _parse_date_from_sheet(sheet_name)

        # 데이터: row 3부터 (index 2)
        for row_idx, row in enumerate(rows[2:], start=3):
            try:
                # 최소 6개 컬럼 필요: 스테이션|히트|종목|참가부|이름|연락처
                if len(row) < 6:
                    continue

                station_raw = row[0]
                heat_raw = row[1]
                event_code_raw = row[2]
                division_raw = row[3]
                name_raw = row[4]
                phone_raw = row[5]

                # 이메일: 6번 또는 7번 컬럼
                email_val = None
                if len(row) > 6 and row[6]:
                    email_val = str(row[6]).strip()
                if not email_val and len(row) > 7 and row[7]:
                    email_val = str(row[7]).strip()

                # 소속
                team_val = None
                # 소속은 이메일 다음 컬럼 (마지막 컬럼)
                last_col_idx = len(row) - 1
                if last_col_idx > 6 and row[last_col_idx]:
                    team_val = str(row[last_col_idx]).strip()

                # 빈 행 스킵
                if station_raw is None or heat_raw is None:
                    continue
                if not event_code_raw or not name_raw:
                    continue

                station = int(station_raw)
                heat_number = int(heat_raw)
                event_code = str(event_code_raw).strip().upper()
                division_str = str(division_raw).strip() if division_raw else "Open"
                phone = _clean_phone(phone_raw)

                if event_code not in EVENT_TYPE_MAP:
                    errors.append(f"시트 '{sheet_name}' {row_idx}행: 알 수 없는 종목코드 '{event_code}'")
                    continue

                # 참가부는 원본 그대로 저장 (매핑은 표시할 때만)
                category = division_str

                # 프로그램 찾기/생성 — division 에 원본 코드 저장
                prog = await _find_or_create_program(
                    db, event_id, event_code, division_str,
                    competition_date, order_counter,
                )
                prog_id_str = str(prog.id)
                if prog_id_str not in program_heats:
                    program_heats[prog_id_str] = {}
                    total_programs += 1

                # 참가자 생성 (팀 종목이면 이름 콤마 분리)
                names_raw = str(name_raw).strip()
                if event_code in TEAM_EVENT_CODES:
                    names = [n.strip() for n in names_raw.replace("，", ",").split(",") if n.strip()]
                else:
                    names = [names_raw]

                participant_ids = []
                participant_names = []
                for pname in names:
                    p = await _find_or_create_participant(
                        db, event_id, pname, phone, email_val, team_val, category,
                    )
                    participant_ids.append(str(p.id))
                    participant_names.append(pname)
                    total_participants += 1

                # 히트 배정 누적
                heat_key = f"{station}_{heat_number}"
                if heat_key not in program_heats[prog_id_str]:
                    program_heats[prog_id_str][heat_key] = {
                        "heat_number": heat_number,
                        "station": station,
                        "participant_ids": [],
                        "participant_names": [],
                    }
                    total_heats += 1
                program_heats[prog_id_str][heat_key]["participant_ids"].extend(participant_ids)
                program_heats[prog_id_str][heat_key]["participant_names"].extend(participant_names)

            except Exception as exc:
                errors.append(f"시트 '{sheet_name}' {row_idx}행: {exc}")

    # 히트 배정 JSON 저장
    for prog_id_str, heats_dict in program_heats.items():
        result = await db.execute(select(Program).where(Program.id == prog_id_str))
        prog = result.scalar_one_or_none()
        if prog:
            existing = prog.heat_assignments or []
            existing_nums = {h["heat_number"] for h in existing if isinstance(h, dict)}
            for hk, hdata in heats_dict.items():
                if hdata["heat_number"] not in existing_nums:
                    existing.append({
                        "heat_number": hdata["heat_number"],
                        "station": hdata["station"],
                        "participant_ids": hdata["participant_ids"],
                        "participant_names": hdata.get("participant_names", []),
                    })
            prog.heat_assignments = existing

    await db.commit()
    wb.close()

    return APIResponse(
        data=ImportSummary(
            programs_created=total_programs,
            participants_imported=total_participants,
            heats_assigned=total_heats,
            sheets_processed=sheets_processed,
            errors=errors[:50],  # 최대 50개
        )
    )
